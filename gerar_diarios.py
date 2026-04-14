#!/usr/bin/env python3
"""
Script para migrar diários existentes para novo formato com suporte a:
- 3 tempos de aula por dia
- Avaliações (mínimo 3)
- Conteúdos por trimestre
- Seleção manual de trimestre
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
from pathlib import Path

# Configurações
PROFESSOR = "LEANDRO CASEMIRO"
TURMAS = {
    "1001": "Matemática",
    "1002": "Matemática",
    "1002 Reforço": "Reforço Escolar de Matemática",
    "1003": "Matemática",
    "2001": "Matemática",
    "2003": "Matemática"
}

DIARIOS_EXISTENTES = {
    "1001": "D:/1. PROJETOS/ESCOLA/2026/DIÁRIOS/Matemática-20260413T205923Z-3-001/Matemática/DIÁRIO 1001 ISMÊNIA 2026.xlsx",
    "1002": "D:/1. PROJETOS/ESCOLA/2026/DIÁRIOS/Matemática-20260413T205923Z-3-001/Matemática/DIÁRIO 1002 ISMÊNIA 2026.xlsx",
    "1002 Reforço": "D:/1. PROJETOS/ESCOLA/2026/DIÁRIOS/Matemática-20260413T205923Z-3-001/Matemática/DIÁRIO 1002 ISMÊNIA 2026 Reforço escolar.xlsx",
    "1003": "D:/1. PROJETOS/ESCOLA/2026/DIÁRIOS/Matemática-20260413T205923Z-3-001/Matemática/DIÁRIO 1003 ISMÊNIA 2026.xlsx",
    "2001": "D:/1. PROJETOS/ESCOLA/2026/DIÁRIOS/Matemática-20260413T205923Z-3-001/Matemática/DIÁRIO 2001 ISMÊNIA 2026.xlsx",
    "2003": "D:/1. PROJETOS/ESCOLA/2026/DIÁRIOS/Matemática-20260413T205923Z-3-001/Matemática/DIÁRIO 2003 ISMÊNIA 2026.xlsx"
}

PASTA_SAIDA = "D:/1. PROJETOS/ESCOLA/2026/DIÁRIOS/app-presenca/diarios"

def extrair_alunos(caminho_arquivo):
    """Extrai lista de alunos da planilha existente"""
    wb = openpyxl.load_workbook(caminho_arquivo, data_only=True)
    ws = wb['1º TRIM']

    alunos = []
    for row in range(4, 200):
        nome = ws.cell(row=row, column=2).value
        if nome and str(nome).strip():
            alunos.append(str(nome).strip())
        else:
            break

    wb.close()
    return alunos

def criar_nova_planilha(turma, disciplina, alunos):
    """Cria nova planilha com estrutura adaptada"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Criar 3 abas de presença (1º, 2º, 3º TRIM)
    for trim_num, trim_nome in [(1, "1º TRIM"), (2, "2º TRIM"), (3, "3º TRIM")]:
        ws = wb.create_sheet(trim_nome)

        # Linha 1: Cabeçalho com info gerais
        ws['A1'] = ""
        ws['B1'] = "DISCIPLINA"
        ws['C1'] = disciplina
        ws['L1'] = "PROFESSOR(A)"
        ws['M1'] = PROFESSOR
        ws['AA1'] = "TURMA"
        ws['AB1'] = turma

        # Linha 2: Meses e Aulas Dadas
        ws['B2'] = "MESES"
        if trim_num == 1:
            ws['C2'] = "FEVEREIRO, MARÇO, ABRIL E MAIO"
        elif trim_num == 2:
            ws['C2'] = "MAIO, JUNHO, JULHO E AGOSTO"
        else:
            ws['C2'] = "SETEMBRO, OUTUBRO, NOVEMBRO E DEZEMBRO"

        ws['AA2'] = "AULAS DADAS"
        ws['AB2'] = 0

        # Linha 3: Cabeçalhos de dados
        ws['A3'] = "Nº"
        ws['B3'] = "NOMES"
        ws['C3'] = "Dias"

        # Colunas de datas (D em diante) - espaço para dados
        # Deixar vazias para serem preenchidas depois
        ws['AA3'] = "TOTAL"

        # Adicionar alunos (linhas 4+)
        for idx, aluno in enumerate(alunos, start=4):
            numero = idx - 3
            ws[f'A{idx}'] = numero
            ws[f'B{idx}'] = aluno
            ws[f'C{idx}'].border = border
            ws[f'AA{idx}'] = 0  # Total de faltas

        # Formatação
        for row in range(1, len(alunos) + 4):
            for col in range(1, 30):  # Cols A-AC
                cell = ws.cell(row=row, column=col)
                if row <= 3:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                else:
                    cell.alignment = center_align
                cell.border = border

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 30
        for col in range(4, 30):
            ws.column_dimensions[get_column_letter(col)].width = 4

    # Criar 3 abas de avaliações (AVAL. E CONT X TRIM)
    for trim_num, trim_nome in [(1, "AVAL. E CONT 1º TRIM"),
                                 (2, "AVAL. E CONT 2º TRIM"),
                                 (3, "AVAL. E CONT 3ºTRIM")]:
        ws = wb.create_sheet(trim_nome)

        # Linha 1: Info gerais
        ws['A1'] = "DISCIPLINA"
        ws['C1'] = disciplina
        ws['L1'] = "PROFESSOR(A)"
        ws['M1'] = PROFESSOR

        # Linha 2: Cabeçalho de avaliações
        ws['A2'] = ""
        ws['B2'] = "NOMES"
        ws['C2'] = f"AVALIAÇÕES {trim_num}º TRIMESTRE"
        ws['L2'] = "TURMA:"
        ws['M2'] = turma

        # Linha 3: Cabeçalhos detalhados
        ws['A3'] = "Nº"
        ws['B3'] = "TIPOS"
        ws['C3'] = "Nº"

        # Colunas D-X: Notas (20 colunas)
        for col_idx in range(4, 24):  # D-W
            col_letter = get_column_letter(col_idx)
            ws[f'{col_letter}3'] = f"AV{col_idx - 3}"

        ws['Y3'] = "TOTAL"
        ws['AA3'] = ""
        ws['AA3'] = f"CONTEÚDO DO {trim_num}º TRIM"

        # Adicionar alunos
        for idx, aluno in enumerate(alunos, start=4):
            numero = idx - 3
            ws[f'A{idx}'] = numero
            ws[f'B{idx}'] = aluno
            ws[f'C{idx}'] = numero
            ws[f'Y{idx}'] = 0  # Total

        # Formatação
        for row in range(1, len(alunos) + 4):
            for col in range(1, 40):
                cell = ws.cell(row=row, column=col)
                if row <= 3:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                else:
                    cell.alignment = center_align
                cell.border = border

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 5
        for col in range(4, 30):
            ws.column_dimensions[get_column_letter(col)].width = 6

    # Criar aba Conselho Final
    ws = wb.create_sheet("CONSELHO FINAL")

    # Cabeçalho
    ws['A1'] = "RESUMO DE APROVEITAMENTO E FREQUÊNCIA TRIMESTRAL"
    ws['C1'] = "ANO LETIVO 2026"
    ws['G1'] = "TURMA"
    ws['H1'] = turma

    ws['C2'] = "PROFESSOR(A)"
    ws['D2'] = PROFESSOR
    ws['I2'] = "DISCIPLINA"
    ws['J2'] = disciplina

    # Linha 3: Trimestres
    ws['D3'] = "1º TRIMESTRE"
    ws['F3'] = "2º TRIMESTRE"
    ws['H3'] = "3º TRIMESTRE"

    # Linha 4: Cabeçalhos
    ws['B4'] = "NOMES"
    ws['C4'] = "NOTAS"
    ws['D4'] = "FALTAS"
    ws['E4'] = "NOTAS"
    ws['F4'] = "FALTAS"
    ws['G4'] = "NOTAS"
    ws['H4'] = "FALTAS"
    ws['I4'] = "REC"
    ws['J4'] = "FINAL"
    ws['K4'] = "RESULTADO"

    # Adicionar alunos
    for idx, aluno in enumerate(alunos, start=5):
        numero = idx - 4
        ws[f'A{idx}'] = numero
        ws[f'B{idx}'] = aluno
        ws[f'K{idx}'] = "reprovado"  # Default

    # Formatação
    for row in range(1, len(alunos) + 5):
        for col in range(1, 15):
            cell = ws.cell(row=row, column=col)
            if row <= 4:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
            else:
                cell.alignment = center_align
            cell.border = border

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30

    return wb

def main():
    """Função principal"""
    # Criar pasta de saída se não existir
    Path(PASTA_SAIDA).mkdir(parents=True, exist_ok=True)

    print("[*] Gerando novos diarios para Leandro Casemiro...")
    print()

    for turma, disciplina in TURMAS.items():
        arquivo_existente = DIARIOS_EXISTENTES[turma]

        if not os.path.exists(arquivo_existente):
            print(f"[!] Arquivo nao encontrado: {arquivo_existente}")
            continue

        print(f"[+] Turma {turma} ({disciplina})...")

        # Extrair alunos
        alunos = extrair_alunos(arquivo_existente)
        print(f"    [OK] {len(alunos)} alunos encontrados")

        # Criar nova planilha
        wb = criar_nova_planilha(turma, disciplina, alunos)

        # Salvar
        nome_saida = f"DIARIO {turma} {disciplina.upper()} 2026.xlsx"
        caminho_saida = os.path.join(PASTA_SAIDA, nome_saida)
        wb.save(caminho_saida)
        wb.close()

        print(f"    [OK] Salvo em: {caminho_saida}")
        print()

    print("[SUCCESS] Todos os diarios foram gerados com sucesso!")
    print(f"[INFO] Localizacao: {PASTA_SAIDA}")

if __name__ == "__main__":
    main()
